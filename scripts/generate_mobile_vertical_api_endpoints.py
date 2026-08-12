#!/usr/bin/env python3
"""Generate mobile vertical API endpoint documentation (xlsx + csv)."""

from __future__ import annotations

import csv
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    import sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = REPO_ROOT / "docs"
XLSX_PATH = OUT_DIR / "mobile-vertical-api-endpoints.xlsx"
CSV_PATH = OUT_DIR / "mobile-vertical-api-endpoints.csv"

COLUMNS = [
    "Vertical",
    "Base URL",
    "HTTP Method",
    "Path",
    "Full URL",
    "Source",
    "Auth",
    "Notes",
]

VERTICALS = {
    "Food Driver": {
        "base": "/api/v1/delivery-mobile",
        "auth": "Delivery mobile auth (JWT)",
        "sheet": "Food Driver",
    },
    "Medical Driver": {
        "base": "/api/v1/medical-mobile/driver",
        "auth": "Medical mobile auth — handler/driver role",
        "sheet": "Medical Driver",
    },
    "Medical Owner": {
        "base": "/api/v1/medical-mobile/owner",
        "auth": "Medical mobile auth — admin + owner persona",
        "sheet": "Medical Owner",
    },
    "Camp Consumer": {
        "base": "/api/v1/camp-consumer",
        "auth": "Camp consumer auth (self-serve registration)",
        "sheet": "Camp Consumer",
    },
}

# (method, path, notes) — extracted from router index.ts files
ROUTES: dict[str, list[tuple[str, str, str]]] = {
    "Food Driver": [
        ("POST", "/auth/login", "Rate-limited auth endpoints"),
        ("POST", "/auth/send-otp", ""),
        ("POST", "/auth/verify-otp", ""),
        ("POST", "/auth/resend-otp", "Same handler as send-otp"),
        ("POST", "/auth/forget-password/otp/send", ""),
        ("POST", "/auth/forget-password/otp/verify", ""),
        ("POST", "/auth/forget-password/set-password", ""),
        ("POST", "/auth/set-password", ""),
        ("POST", "/auth/forget-password/otp/resend", "Same handler as forget-password send"),
        ("POST", "/auth/reset-password", ""),
        ("POST", "/auth/check-account", ""),
        ("POST", "/auth/logout", ""),
        ("POST", "/auth/refresh", ""),
        ("GET", "/account/me", ""),
        ("GET", "/profile", "Alias for account/me"),
        ("PUT", "/account", "Triggers OTP confirm flow via separate endpoints"),
        ("PATCH", "/account/update/resend-otp", "Sensitive OTP rate limit"),
        ("PATCH", "/account/confirm", "Sensitive OTP rate limit"),
        ("POST", "/account/transfer-ownership", ""),
        ("POST", "/account/transfer-ownership/verify", "Sensitive OTP rate limit"),
        ("GET", "/account/mygrubpacs", ""),
        ("DELETE", "/account", ""),
        ("GET", "/dashboard", ""),
        ("GET", "/support/category", ""),
        ("GET", "/support/faq", ""),
        ("GET", "/support/answer", ""),
        ("POST", "/employee", "Manager-only employee management"),
        ("GET", "/employee", ""),
        ("GET", "/employee/dropdowns", ""),
        ("PATCH", "/employee/suspend", ""),
        ("PATCH", "/employee/reactivate", ""),
        ("DELETE", "/employee", ""),
        ("POST", "/restaurant", "Manager-only restaurant management"),
        ("GET", "/restaurant", ""),
        ("PATCH", "/restaurant/resource/unassign", ""),
        ("PATCH", "/restaurant/resource/suspend", ""),
        ("GET", "/restaurant/:id", ""),
        ("PUT", "/restaurant/:id", ""),
        ("GET", "/config", "Public config (no auth)"),
        ("GET", "/boxes", ""),
        ("POST", "/boxes", "QR register / add box"),
        ("GET", "/boxes/:box_id", ""),
        ("DELETE", "/boxes/:box_id", ""),
        ("PATCH", "/boxes/:box_id/settings", ""),
        ("POST", "/boxes/:box_id/connection", ""),
        ("DELETE", "/boxes/:box_id/connection", ""),
        ("POST", "/boxes/:box_id/lock/otp", "GrubLock OTP request; sensitive OTP rate limit"),
        ("POST", "/boxes/:box_id/lock/verify", "GrubLock OTP verify + unlock; sensitive OTP rate limit"),
        ("PATCH", "/boxes/:box_id/lock", "GrubLock lock/unlock toggle after OTP verify"),
        ("GET", "/notification", ""),
        ("PATCH", "/notification", "Mark read / dismiss"),
        ("POST", "/notification/test-trigger", "Dev/test only"),
    ],
    "Medical Driver": [
        ("GET", "/health", "Vertical health probe"),
        ("POST", "/auth/login", "Handler role only"),
        ("POST", "/auth/send-otp", ""),
        ("POST", "/auth/verify-otp", ""),
        ("POST", "/auth/resend-otp", ""),
        ("POST", "/auth/forget-password/otp/send", ""),
        ("POST", "/auth/forget-password/otp/verify", ""),
        ("POST", "/auth/forget-password/set-password", ""),
        ("POST", "/auth/set-password", ""),
        ("POST", "/auth/forget-password/otp/resend", ""),
        ("POST", "/auth/reset-password", ""),
        ("POST", "/auth/check-account", ""),
        ("POST", "/auth/logout", ""),
        ("POST", "/auth/refresh", ""),
        ("GET", "/account/me", ""),
        ("GET", "/profile", ""),
        ("PUT", "/account/password", "In-app password change"),
        ("DELETE", "/account", ""),
        ("GET", "/dashboard", ""),
        ("GET", "/support/category", ""),
        ("GET", "/support/faq", ""),
        ("GET", "/support/answer", ""),
        ("GET", "/config", ""),
        ("GET", "/boxes", ""),
        ("POST", "/boxes", "QR register (same as Food Driver)"),
        ("GET", "/boxes/:box_id", ""),
        ("DELETE", "/boxes/:box_id", ""),
        ("PATCH", "/boxes/:box_id/settings", ""),
        ("POST", "/boxes/:box_id/connection", ""),
        ("DELETE", "/boxes/:box_id/connection", ""),
        ("POST", "/boxes/:box_id/lock/otp", "GrubLock OTP unlock flow"),
        ("POST", "/boxes/:box_id/lock/verify", ""),
        ("PATCH", "/boxes/:box_id/lock", ""),
        ("GET", "/boxes/:box_id/location", "Medical-only: box GPS location"),
        ("POST", "/boxes/:box_id/location/share", "Medical-only: share location link"),
        ("GET", "/boxes/:box_id/diagnostics", "Medical-only: device diagnostics"),
        ("GET", "/boxes/:box_id/alerts", "Medical-only: temperature/telemetry alerts"),
        ("GET", "/emergency/call-metadata", "Medical-only: emergency call UI metadata"),
        ("POST", "/emergency/alert", "Medical-only: post emergency alert"),
        ("GET", "/notification", ""),
        ("PATCH", "/notification", ""),
    ],
    "Medical Owner": [
        ("GET", "/health", "Vertical health probe"),
        ("POST", "/auth/login", "Admin role + owner persona only"),
        ("POST", "/auth/send-otp", ""),
        ("POST", "/auth/verify-otp", ""),
        ("POST", "/auth/resend-otp", ""),
        ("POST", "/auth/forget-password/otp/send", ""),
        ("POST", "/auth/forget-password/otp/verify", ""),
        ("POST", "/auth/forget-password/set-password", ""),
        ("POST", "/auth/set-password", ""),
        ("POST", "/auth/forget-password/otp/resend", ""),
        ("POST", "/auth/reset-password", ""),
        ("POST", "/auth/check-account", ""),
        ("POST", "/auth/logout", ""),
        ("POST", "/auth/refresh", ""),
        ("GET", "/account/me", ""),
        ("GET", "/profile", ""),
        ("PUT", "/account", "Owner profile update (no OTP confirm flow)"),
        ("PUT", "/account/password", "In-app password change"),
        ("DELETE", "/account", ""),
        ("GET", "/dashboard", ""),
        ("GET", "/support/category", ""),
        ("GET", "/support/faq", ""),
        ("GET", "/support/answer", ""),
        ("GET", "/config", ""),
        ("GET", "/boxes", ""),
        ("POST", "/boxes/claim", "Owner-only: claim pre-assigned box (no QR register)"),
        ("GET", "/boxes/:box_id", ""),
        ("DELETE", "/boxes/:box_id", ""),
        ("PATCH", "/boxes/:box_id/settings", ""),
        ("POST", "/boxes/:box_id/connection", ""),
        ("DELETE", "/boxes/:box_id/connection", ""),
        ("GET", "/boxes/:box_id/location", ""),
        ("POST", "/boxes/:box_id/location/share", "Owner: share box location link"),
        ("GET", "/boxes/:box_id/diagnostics", "Owner: hardware diagnostics accordion"),
        ("GET", "/boxes/:box_id/alerts", "Owner: box-scoped alerts with filters"),
        ("GET", "/emergency/call-metadata", "Facility phone from client record"),
        ("PATCH", "/boxes/:box_id/lock", "Direct lock/unlock — no OTP flow"),
        ("PATCH", "/grublock/lock", "Owner-only: lock GrubLock without OTP"),
        ("PATCH", "/grublock/emergency_unlock", "Owner-only: emergency unlock without OTP"),
        ("GET", "/notification", ""),
        ("PATCH", "/notification", ""),
    ],
    "Camp Consumer": [
        ("GET", "/health", "Vertical health probe"),
        ("POST", "/auth/login", "Self-serve consumer registration/login"),
        ("POST", "/auth/send-otp", ""),
        ("POST", "/auth/verify-otp", ""),
        ("POST", "/auth/resend-otp", ""),
        ("POST", "/auth/forget-password/otp/send", ""),
        ("POST", "/auth/forget-password/otp/verify", ""),
        ("POST", "/auth/forget-password/set-password", ""),
        ("POST", "/auth/set-password", ""),
        ("POST", "/auth/forget-password/otp/resend", ""),
        ("POST", "/auth/reset-password", ""),
        ("POST", "/auth/check-account", ""),
        ("POST", "/auth/logout", ""),
        ("POST", "/auth/refresh", ""),
        ("GET", "/account/me", ""),
        ("GET", "/profile", ""),
        ("PUT", "/account", "Consumer profile update"),
        ("PUT", "/account/password", "Camp-only: change password in-app"),
        ("DELETE", "/account", ""),
        ("GET", "/dashboard", ""),
        ("GET", "/support/category", ""),
        ("GET", "/support/faq", ""),
        ("GET", "/support/answer", ""),
        ("GET", "/config", ""),
        ("GET", "/boxes", ""),
        ("POST", "/boxes", "QR register"),
        ("GET", "/boxes/:box_id", ""),
        ("DELETE", "/boxes/:box_id", ""),
        ("PATCH", "/boxes/:box_id/settings", ""),
        ("POST", "/boxes/:box_id/connection", ""),
        ("DELETE", "/boxes/:box_id/connection", ""),
        ("POST", "/boxes/:box_id/lock/otp", "GrubLock OTP unlock flow"),
        ("POST", "/boxes/:box_id/lock/verify", ""),
        ("PATCH", "/boxes/:box_id/lock", ""),
        ("GET", "/boxes/:box_id/alerts", "Camp-only: alert thresholds/status"),
        ("PATCH", "/boxes/:box_id/alerts", "Camp-only: configure alert settings"),
        ("GET", "/boxes/:box_id/camera/live", "Camp-only: live camera HLS (client-scoped S3)"),
        ("GET", "/boxes/:box_id/camera/feeds", "Camp-only: list camera feeds"),
        ("GET", "/boxes/:box_id/camera/feeds/:feed_id/stream", "Camp-only: stream URL/token"),
        ("POST", "/boxes/:box_id/camera/upload-url", "Camp-only: presigned PUT for box/device ingest"),
        ("POST", "/boxes/:box_id/camera/feeds/register", "Camp-only: register recording after S3 upload"),
        ("PATCH", "/boxes/:box_id/camera/surveillance-mode", "Camp-only: toggle surveillance mode"),
        ("GET", "/notification", ""),
        ("PATCH", "/notification", ""),
    ],
}

# Delivery routes that are intentionally absent on med/camp verticals
BLOCKED_ON_NEW_VERTICALS: list[tuple[str, str, str]] = [
    ("PUT", "/account", "No account update flow on Medical Driver"),
    ("PATCH", "/account/update/resend-otp", "Blocked on all new verticals"),
    ("PATCH", "/account/confirm", "Blocked on all new verticals"),
    ("POST", "/account/transfer-ownership", "Blocked on all new verticals"),
    ("POST", "/account/transfer-ownership/verify", "Blocked on all new verticals"),
    ("GET", "/account/mygrubpacs", "Blocked on all new verticals"),
    ("POST", "/employee", "Manager feature — not in med/camp apps"),
    ("GET", "/employee", ""),
    ("GET", "/employee/dropdowns", ""),
    ("PATCH", "/employee/suspend", ""),
    ("PATCH", "/employee/reactivate", ""),
    ("DELETE", "/employee", ""),
    ("POST", "/restaurant", "Manager feature — not in med/camp apps"),
    ("GET", "/restaurant", ""),
    ("PATCH", "/restaurant/resource/unassign", ""),
    ("PATCH", "/restaurant/resource/suspend", ""),
    ("GET", "/restaurant/:id", ""),
    ("PUT", "/restaurant/:id", ""),
    ("POST", "/notification/test-trigger", "Dev/test — not exposed on new verticals"),
]

# Owner-specific blocks vs delivery baseline
OWNER_BLOCKED_EXTRA: list[tuple[str, str, str]] = [
    ("POST", "/boxes", "Owner uses POST /boxes/claim instead of QR register"),
    ("POST", "/boxes/:box_id/lock/otp", "Owner: no OTP lock flow"),
    ("POST", "/boxes/:box_id/lock/verify", "Owner: no OTP lock flow"),
]


def route_key(method: str, path: str) -> str:
    return f"{method.upper()} {path}"


def delivery_patterns() -> set[str]:
    return {route_key(m, p) for m, p, _ in ROUTES["Food Driver"]}


def classify_source(vertical: str, method: str, path: str) -> str:
    key = route_key(method, path)
    if vertical == "Food Driver":
        return "Baseline (Food Driver)"
    if key in delivery_patterns():
        return "Reused from delivery-mobile"
    return "New — vertical-specific"


def make_row(
    vertical: str,
    method: str,
    path: str,
    notes: str,
    *,
    source_override: str | None = None,
) -> dict[str, str]:
    meta = VERTICALS[vertical]
    base = meta["base"]
    return {
        "Vertical": vertical,
        "Base URL": base,
        "HTTP Method": method.upper(),
        "Path": path,
        "Full URL": f"{base}{path}",
        "Source": source_override or classify_source(vertical, method, path),
        "Auth": meta["auth"],
        "Notes": notes,
    }


def vertical_sheet_rows(vertical: str) -> list[dict[str, str]]:
    rows = [
        make_row(vertical, method, path, notes)
        for method, path, notes in ROUTES[vertical]
    ]

    if vertical == "Food Driver":
        return rows

    existing = {route_key(r["HTTP Method"], r["Path"]) for r in rows}

    blocked = list(BLOCKED_ON_NEW_VERTICALS)
    if vertical == "Medical Owner":
        blocked = blocked + OWNER_BLOCKED_EXTRA
    if vertical == "Medical Driver":
        blocked = [(m, p, n) for m, p, n in blocked if not (m == "PUT" and p == "/account")]

    for method, path, notes in blocked:
        key = route_key(method, path)
        if key in existing:
            continue
        rows.append(
            make_row(
                vertical,
                method,
                path,
                notes or "Exists on Food Driver but not mounted on this vertical",
                source_override="Blocked on this vertical",
            )
        )

    rows.sort(key=lambda r: (r["Path"], r["HTTP Method"]))
    return rows


def reused_rows() -> list[dict[str, str]]:
    baseline = delivery_patterns()
    rows: list[dict[str, str]] = []
    for vertical in ("Medical Driver", "Medical Owner", "Camp Consumer"):
        for method, path, notes in ROUTES[vertical]:
            if route_key(method, path) in baseline:
                row = make_row(vertical, method, path, notes)
                row["Source"] = "Reused from delivery-mobile"
                rows.append(row)
    rows.sort(key=lambda r: (r["Path"], r["HTTP Method"], r["Vertical"]))
    return rows


def comparison_matrix_rows() -> list[dict[str, str]]:
    all_routes: dict[str, tuple[str, str, str]] = {}

    def add(method: str, path: str, note: str = "") -> None:
        key = route_key(method, path)
        if key not in all_routes:
            all_routes[key] = (method, path, note)

    for _v, routes in ROUTES.items():
        for method, path, notes in routes:
            add(method, path, notes)

    for method, path, notes in BLOCKED_ON_NEW_VERTICALS + OWNER_BLOCKED_EXTRA:
        add(method, path, notes)

    def status(vertical: str, method: str, path: str) -> str:
        key = route_key(method, path)
        mounted = key in {route_key(m, p) for m, p, _ in ROUTES[vertical]}

        if vertical == "Food Driver":
            return "✅" if mounted else "❌"

        if mounted:
            if key in delivery_patterns():
                return "✅ reused"
            return "🆕 new"

        blocked_keys = {route_key(m, p) for m, p, _ in BLOCKED_ON_NEW_VERTICALS}
        if vertical == "Medical Owner":
            blocked_keys |= {route_key(m, p) for m, p, _ in OWNER_BLOCKED_EXTRA}
        if vertical == "Medical Driver" and method == "PUT" and path == "/account":
            blocked_keys.add(key)

        if key in delivery_patterns() and key in blocked_keys:
            return "❌ blocked"
        if key in delivery_patterns():
            return "❌"
        return "—"

    rows: list[dict[str, str]] = []
    for method, path, notes in sorted(all_routes.values(), key=lambda x: (x[1], x[0])):
        rows.append(
            {
                "HTTP Method": method,
                "Path": path,
                "Food Driver": status("Food Driver", method, path),
                "Medical Driver": status("Medical Driver", method, path),
                "Medical Owner": status("Medical Owner", method, path),
                "Camp Consumer": status("Camp Consumer", method, path),
                "Notes": notes,
            }
        )
    return rows


def autosize_columns(ws, max_width: int = 60) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 12), max_width)


def write_sheet_header(ws, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def write_rows(ws, headers: list[str], rows: list[dict[str, str]]) -> None:
    write_sheet_header(ws, headers)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
    autosize_columns(ws)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    all_endpoint_rows: list[dict[str, str]] = []
    sheet_counts: dict[str, int] = {}

    wb = Workbook()
    wb.remove(wb.active)

    for vertical in ("Medical Driver", "Medical Owner", "Camp Consumer", "Food Driver"):
        sheet_name = VERTICALS[vertical]["sheet"]
        rows = vertical_sheet_rows(vertical)
        ws = wb.create_sheet(sheet_name)
        write_rows(ws, COLUMNS, rows)
        sheet_counts[sheet_name] = len(rows)
        all_endpoint_rows.extend(rows)

    comp_headers = [
        "HTTP Method",
        "Path",
        "Food Driver",
        "Medical Driver",
        "Medical Owner",
        "Camp Consumer",
        "Notes",
    ]
    comp_rows = comparison_matrix_rows()
    ws_comp = wb.create_sheet("Comparison Matrix")
    write_rows(ws_comp, comp_headers, comp_rows)
    sheet_counts["Comparison Matrix"] = len(comp_rows)

    reused = reused_rows()
    ws_reused = wb.create_sheet("Reused from Food Driver")
    write_rows(ws_reused, COLUMNS, reused)
    sheet_counts["Reused from Food Driver"] = len(reused)

    wb.save(XLSX_PATH)

    with CSV_PATH.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(all_endpoint_rows)

    print(f"Wrote {XLSX_PATH}")
    print(f"Wrote {CSV_PATH}")
    print("Row counts per sheet:")
    for name, count in sheet_counts.items():
        print(f"  {name}: {count}")
    print(f"  CSV (all vertical endpoint rows): {len(all_endpoint_rows)}")


if __name__ == "__main__":
    main()
